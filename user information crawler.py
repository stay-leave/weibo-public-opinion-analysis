#coding='utf-8'
import xlrd
import re
import requests
import xlwt
import os
import time as t
import random
import numpy as np	
import datetime
import urllib3
import sys
from multiprocessing.dummy import Pool as ThreadPool
import openpyxl

urllib3.disable_warnings()
cookie=''

headers = {
				  'Accept-Encoding': 'gzip, deflate, br',
				 'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
				 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0',
				'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
				 'Referer': 'https://weibo.cn/',
			    'Connection': 'keep-alive',
				'Cookie': cookie,
				}

def save_file(alls,name):
    """将数据保存在一个excel
    """
    f=openpyxl.Workbook() 
    sheet1=f.create_sheet('sheet1')
    sheet1['A1']='uid'
    sheet1['B1']='昵称'
    sheet1['C1']='性别'
    sheet1['D1']='地区'
    sheet1['E1']='生日'

    i=2#openpyxl最小值是1，写入的是xlsx
    for all in alls:#遍历每一页
        #for data in all:#遍历每一行
            for j in range(1,len(all)+1):#取每一单元格
                #sheet1.write(i,j,all[j])#写入单元格
                sheet1.cell(row=i,column=j,value=all[j-1])
            i=i+1#往下一行

    f.save(str(name))

def extract(inpath,l):
    """取出一列数据"""
    data = xlrd.open_workbook(inpath, encoding_override='utf-8')
    table = data.sheets()[0]#选定表
    nrows = table.nrows#获取行号
    ncols = table.ncols#获取列号
    numbers=[]
    for i in range(1, nrows):#第0行为表头
        alldata = table.row_values(i)#循环输出excel表中每一行，即所有数据
        result = alldata[l]#取出表中第一列数据
        numbers.append(result)
    return numbers

def require(url):
	"""获取网页源码"""
	while True:
		try:
			response = requests.get(url, headers=headers,timeout=(30,50),verify=False)
			#print(url)
			code_1=response.status_code
			#print(type(code_1))
			#t.sleep(random.randint(1,2))
			if code_1==200:
				print('正常爬取中，状态码：'+str(code_1))#状态码
				t.sleep(random.randint(1,2))
				break
			else:
				print('请求异常，重试中，状态码为：'+str(code_1))#状态码
				t.sleep(random.randint(2,3))
				continue
		except:
			t.sleep(random.randint(2,3))
			continue

	#print(response.encoding)#首选编码
	#response.encoding=response.apparent_encoding
	html=response.text#源代码文本
	return html

def body(html):
	"""单个资料爬取"""
	data=re.findall('<div class="tip">基本信息</div>(.*?)<div class="tip">其他信息</div>',html,re.S)#取大
	#print(data)
	name_0=re.findall('<div class="c">昵称:(.*?)<br/>',str(data),re.S)#用户昵称
	#print(name_0)
	try:
		name_1=re.findall('<br/>性别:(.*?)<br/>',str(data),re.S)#性别
	except:
		name_1=['无']
	try:
		name_2=re.findall('<br/>地区:(.*?)<br/>',str(data),re.S)#地区
	except:
		name_2=['无']
	try:
		name_3=re.findall('<br/>生日:(\d{4}-\d{1,2}-\d{1,2})<br/>',str(data),re.S)#生日
	except:
		name_3=['无']
	all=name_0+name_1+name_2+name_3
	return all

def run(uid):
	uid=int(uid)
	alls=[]
	url='https://weibo.cn/'+str(uid)+'/info'
	one_data=[uid]+body(require(url))
	#t.sleep(1)
	alls.append(one_data)
	return alls


if __name__ == '__main__':
	#start_time = t.clock()
	uids=list(set(extract(r'2.数据处理\评论信息.xlsx',0)))
	#print(len(uids))
	pool = ThreadPool()
	alls_1=pool.map(run, uids)
	#print(len(alls_1))
	alls_2=[]
	for i in alls_1:
		alls_2.append(i[0])
	#print(len(alls_2))
	save_file(alls_2,'我.xlsx')#保存路径
	#stop_time = t.clock()
	#cost = stop_time - start_time
	#print("%s cost %s second" % (os.path.basename(sys.argv[0]), cost))

		
		


