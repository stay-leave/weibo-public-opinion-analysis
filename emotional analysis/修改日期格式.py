#coding='utf-8'
import xlrd
import re
import xlwt
import time as t
import openpyxl

def extract_1(inpath):
    """提取评论数据"""
    data = xlrd.open_workbook(inpath, encoding_override='utf-8')
    table = data.sheets()[0]#选定表
    nrows = table.nrows#获取行号
    ncols = table.ncols#获取列号
    numbers=[]
    for i in range(1, nrows):#第0行为表头
        alldata = table.row_values(i)#循环输出excel表中每一行，即所有数据
        result_0 = int(alldata[0])#uid
        result_1 = alldata[1]#用户名
        result_2 = alldata[2]#性别
        result_3 = alldata[3]#地区
        result_4 = alldata[4]#生日
        result_5 = alldata[5]#评论
        result_6 = alldata[6]#日期
        result_7 = alldata[7]#极性
        result_8 = alldata[8]#情感值
        result_9 = alldata[9]#情感值
        numbers.append([result_0,result_1,result_2,result_3,result_4,result_5,result_6,result_7,result_8,result_9])
    return numbers



def save_file(alls,name):
    """将一个时间段的所有评论数据保存在一个excel
    """
    f=openpyxl.Workbook() 
    sheet1=f.create_sheet('sheet1')
    sheet1['A1']='uid'
    sheet1['B1']='昵称'
    sheet1['C1']='性别'
    sheet1['D1']='省份'
    sheet1['E1']='城市'
    sheet1['F1']='生日'
    sheet1['G1']='评论'
    sheet1['H1']='日期'
    sheet1['I1']='情感极性'
    sheet1['J1']='情感值'

    i=2#openpyxl最小值是1，写入的是xlsx
    for all in alls:#遍历每一页
        #for data in all:#遍历每一行
            for j in range(1,len(all)+1):#取每一单元格
                #sheet1.write(i,j,all[j])#写入单元格
                sheet1.cell(row=i,column=j,value=all[j-1])
            i=i+1#往下一行

    f.save(str(name))

if __name__ == "__main__":
    h_1=extract_1('合集.xlsx')
    h=[]
    for i in h_1:
        if len(i[7])==6:
            i[7]=i[7].replace('月','')
            i[7]=i[7].replace('日','')
            i[7]='2021-'+i[7][0:2]+'-'+i[7][2:4]
        h.append([i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9]])
    save_file(h,'合集_日期统一.xlsx')

    
    
   

        
