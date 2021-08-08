# coding=utf-8
import requests
import json
import xlrd
import re
import xlwt
import time as t
import openpyxl

""" 
你的 APPID AK SK 
每秒钟只能调用两次
"""
APP_ID = ''
API_KEY = ''
SECRET_KEY = ''

# client = AipNlp(APP_ID, API_KEY, SECRET_KEY)

# 获取token
# client_id 为官网获取的AK， client_secret 为官网获取的SK
host = 'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=' + API_KEY + '&client_secret=' + SECRET_KEY
response = requests.get(host)
while True:
    if response.status_code == 200:
        info = json.loads(response.text)  # 将字符串转成字典
        access_token = info['access_token']  # 解析数据到access_token
        break
    else:
        continue

access_token=access_token

def extract(inpath):
    """提取数据"""
    data = xlrd.open_workbook(inpath, encoding_override='utf-8')
    table = data.sheets()[0]  # 选定表
    nrows = table.nrows  # 获取行号
    ncols = table.ncols  # 获取列号
    numbers = []
    for i in range(1, nrows):  # 第0行为表头
        alldata = table.row_values(i)  # 循环输出excel表中每一行，即所有数据
        result_0 = alldata[0]  # 取出id
        result_1 = alldata[1]  # 取出网友名
        result_2 = alldata[2]  # 取出网友评论
        result_3 = alldata[3]  # 取出日期
        numbers.append([result_0, result_1, result_2, result_3])
    return numbers


def emotion(text):
    while True:  # 处理aps并发异常
        url = 'https://aip.baidubce.com/rpc/2.0/nlp/v1/sentiment_classify_custom?charset=UTF-8&access_token=' + access_token
        #headers = {'Content-Type': 'application/json', 'Connection': 'close'}  # headers=headers
        body = {'text': text[:1024]}
        requests.packages.urllib3.disable_warnings()
        try:
            res = requests.post(url=url, data=json.dumps(body), verify=False)
            rc=res.status_code
            print(rc)
        except:
            print('发生错误，停五秒重试')
            t.sleep(5)
            continue
        if rc==200:
            print('正常请求中')
        else:
            print('遇到错误，重试')
            t.sleep(2)
            continue
        try:
            judge = res.text
            print(judge)
        except:
            print('错误,正在重试，错误文本为：' + text)
            t.sleep(1)
            continue
        if judge == {'error_code': 18, 'error_msg': 'Open api qps request limit reached'}:
            print('并发量限制')
            t.sleep(1)
            continue
        elif 'error_msg' in judge:  # 如果出现意外的报错，就结束本次循环
            print('其他错误')
            t.sleep(1)
            continue
        else:
            break
    # print(judge)
    judge=eval(judge)#将字符串转换为字典
    #print(type(judge))
    pm = judge["items"][0]["sentiment"]  # 情感分类
    #print(pm)
    if pm == 0:
        pm = '负向'
    elif pm == 1:
        pm = '中性'
    else:
        pm = '正向'
    pp = judge["items"][0]["positive_prob"]  # 正向概率
    pp = round(pp, 4)
    #print(pp)
    np = judge["items"][0]["negative_prob"]  # 负向概率
    np = round(np, 4)
    #print(np)
    return pm, pp, np


def run(inpath):
    "运行程序,返回一个嵌套小列表的大列表"
    alls = []
    all = extract(inpath)
    for i in all:
        id = i[0]
        name = i[1]
        review = i[2]  # 网友评论
        # review= emotion(review)
        date = i[3]
        pm, pp, np = emotion(review)
        alls.append([id, name, review, date, pm, pp])  # 只取正向，将所有放置在一个区间
        t.sleep(1)
    return alls


def save_file(alls, name):
    """将一个时间段的所有评论数据保存在一个excle
    """
    f = openpyxl.Workbook()
    sheet1 = f.create_sheet('sheet1')
    sheet1['A1'] = 'uid'
    sheet1['B1'] = '昵称'
    sheet1['C1'] = '评论内容'
    sheet1['D1'] = '日期'
    sheet1['E1'] = '评论情感极性'
    sheet1['F1'] = '评论情感概率'  # [0,0.5]负向，(0.5,1]正向

    i = 2  # openpyxl最小值是1，写入的是xlsx
    for all in alls:  # 遍历每一页
        # for data in all:#遍历每一行
        for j in range(1, len(all) + 1):  # 取每一单元格
            # sheet1.write(i,j,all[j])#写入单元格
            sheet1.cell(row=i, column=j, value=all[j - 1])
        i = i + 1  # 往下一行

    f.save(str(name))


if __name__ == "__main__":
    save_file(run( '三.xlsx'),  '三情感值.xlsx')
